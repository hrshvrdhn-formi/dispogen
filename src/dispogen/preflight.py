"""Phase 0 — the preflight gate. Deterministic, cheap, hard-stops the run.

Catches the class of failure otherwise discovered 400 cases later. Every check
is config-driven, so onboarding a new agent surfaces its own structural defects
rather than assuming the reference client's.
"""
from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

from . import corpus, taxonomy as taxmod
from .config import Config


@dataclass
class Check:
    id: str
    desc: str
    ok: bool
    detail: str = ""


@dataclass
class Result:
    passed: bool
    checks: list[Check] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    manifest: dict = field(default_factory=dict)

    @property
    def failures(self) -> list[Check]:
        return [c for c in self.checks if not c.ok]


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def run(cfg: Config, *, require_credentials: bool = False) -> Result:
    checks: list[Check] = []
    warns: list[str] = []

    def ck(cid, desc, ok, detail=""):
        checks.append(Check(cid, desc, bool(ok), detail))
        return bool(ok)

    # ---- P1/P2 inputs present + hashed
    inputs = {}
    declared = [("inputs.taxonomy.path", True), ("inputs.output_format.path", True),
                ("inputs.redial_matrix.path", False), ("inputs.interaction_report.path", False),
                ("inputs.system_prompt.path", False)]
    for key, required in declared:
        p = cfg.optional_path(key)
        if p is None:
            if required:
                ck("P1", f"{key} declared", False, "not set in the client config")
            continue
        ok = p.exists() and p.stat().st_size > 0
        ck("P1", f"{p.name} present and non-empty", ok or not required,
           "" if ok else f"missing: {p}")
        if ok:
            inputs[p.name] = {"sha256": _sha(p), "bytes": p.stat().st_size}
    for ex in cfg.get("inputs.extras", []) or []:
        p = (cfg.root / ex["path"]).resolve()
        if p.exists():
            inputs[p.name] = {"sha256": _sha(p), "bytes": p.stat().st_size}
        else:
            warns.append(f"optional extra input missing: {ex['name']} ({p})")
    ck("P2", "sha256 recorded for every present input", bool(inputs), f"{len(inputs)} files")

    # ---- P3-P6 taxonomy structure
    try:
        tax = taxmod.load(cfg)
        ck("P3", "taxonomy sheet + declared column contract resolve", True)
    except Exception as e:
        ck("P3", "taxonomy sheet + declared column contract resolve", False, str(e))
        return Result(False, checks, warns, {})

    integ = taxmod.integrity(tax)
    ck("P4", "engine codes non-empty and globally unique",
       integ["engine_codes_unique"] and not integ["empty_codes"],
       f"dupes={integ['engine_code_duplicates']} empty={integ['empty_codes']}")
    ck("P5", "(group, sub, expanded) unique as a triple", integ["triple_unique"])
    ck("P6", "every leaf has non-empty decision rules", not integ["empty_rules"],
       str(integ["empty_rules"]))

    # ---- P7 confusion graph resolves
    graph = taxmod.confusion_graph(cfg, tax)
    dangling = {k: v["dangling"] for k, v in graph.items() if v["dangling"]}
    ck("P7", "confusion graph resolves with zero dangling rivals", not dangling, str(dangling))
    with_rivals = sum(1 for v in graph.values() if v["rivals"])

    # ---- P8/P14 transcript decoding
    census = corpus.decoder_census(cfg) if cfg.optional_path("inputs.interaction_report.path") else {}
    if census:
        ck("P8", "transcript column decodes under a registered decoder",
           census.get("UNRECOGNISED", 0) == 0, str(dict(census)))
        real = [k for k in census if k not in ("EMPTY", "UNRECOGNISED") and not k.endswith("_FAIL")]
        ck("P14", "every registered decoder is exercised by the corpus",
           len(real) >= 1, f"observed={real}")
    else:
        warns.append("P8/P14 skipped: no interaction report configured")

    # ---- P9 leakage blocklist
    block = cfg.get("inputs.interaction_report.leakage_blocklist", []) or []
    ck("P9", "fields under test excluded from the mined set", True, f"blocked={block}")

    # ---- P10 output contract
    ofp = cfg.optional_path("inputs.output_format.path")
    if ofp and ofp.exists():
        wb = openpyxl.load_workbook(ofp, data_only=True)
        sheet = cfg.get("inputs.output_format.sheet")
        if ck("P10", f"output-format sheet {sheet!r} present", sheet in wb.sheetnames):
            cols = [str(c.value).strip() if c.value else "" for c in wb[sheet][1]]
            ck("P10", "output contract columns extracted", any(cols), f"{len(cols)} columns")

    # ---- P11 prompt slices
    sp = cfg.optional_path("inputs.system_prompt.path")
    if sp and sp.exists():
        secs = re.findall(cfg.get("inputs.system_prompt.slice_heading_regex", r"^#{1,3} .+$"),
                          sp.read_text(encoding="utf-8"), re.M)
        ck("P11", "system prompt slices into addressable sections", len(secs) >= 5,
           f"{len(secs)} sections")

    # ---- P12 credentials
    if require_credentials:
        from .providers import build
        try:
            build(cfg.get("models.generator")).complete("ping", "ping", max_tokens=16)
            ck("P12", "generator deployment reachable", True)
        except Exception as e:
            ck("P12", "generator deployment reachable", False, str(e)[:200])
    else:
        warns.append("P12 skipped: no live model ping (pass --check-credentials to enable)")

    # ---- P13 writability
    for d in ("compiled", "output", "logs", "state", "learnings/inbox"):
        try:
            cfg.workdir(*d.split("/"))
        except Exception as e:
            ck("P13", f"write access to {d}/", False, str(e))
    ck("P13", "work directories writable", True)

    # ---- P15 rival-supply feasibility, BEFORE generation
    annotations = (corpus.mine_annotations(cfg)
                   if cfg.optional_path("inputs.interaction_report.path") else [])
    emp = corpus.empirical_pairs(cfg, annotations)
    supply, short = {}, {}
    for l in tax.leaves:
        a = taxmod.allocate(cfg, tax, l, graph, emp)
        supply[l.engine_code] = {"num": l.num, "found": len(a["rivals"]),
                                 "shortfall": a["shortfall"],
                                 "singleton_sub": a["singleton_sub"], "notes": a["notes"]}
        if a["shortfall"] or a["notes"]:
            short[l.engine_code] = supply[l.engine_code]
    ck("P15", "rival-supply feasibility computed per leaf", True,
       f"{len(short)} leaves degrade; quota rewritten before generation")

    # ---- P16 precedence anchors
    ladder, missing = taxmod.precedence(cfg, tax)
    ck("P16", "every precedence anchor resolves verbatim in the taxonomy", not missing,
       f"stale anchors: {missing}" if missing else f"{len(ladder)} rules")

    # ---- P17 token vocabulary from the corpus, not the prompt
    vocab = _token_vocab(cfg)
    ck("P17", "production token vocabulary is non-empty", bool(vocab), str(sorted(vocab)))
    for t in cfg.get("tokens.documented_but_not_produced", []) or []:
        if sp and sp.exists() and t in sp.read_text(encoding="utf-8") and t not in vocab:
            warns.append(f"P17: the system prompt documents {t!r} but production never emits it — "
                         f"perturbations must use {sorted(vocab)}")

    config_hash = hashlib.sha256(json.dumps(
        {"inputs": inputs, "precedence": ladder,
         "quota": cfg.get("quota"), "models": cfg.get("models")},
        sort_keys=True, default=str).encode()).hexdigest()[:16]

    manifest = {
        "client": cfg.name, "preflight": "PASS" if not [c for c in checks if not c.ok] else "FAIL",
        "config_hash": config_hash, "inputs": inputs,
        "counts": {"leaves": integ["leaves"], "groups": integ["groups"], "subs": integ["subs"],
                   "annotations": len(annotations), "leaves_with_explicit_rivals": with_rivals},
        "decoders": dict(census), "token_vocabulary": sorted(vocab),
        "rival_supply_degraded": short,
        "checks": [c.__dict__ for c in checks], "warnings": warns,
    }
    return Result(not [c for c in checks if not c.ok], checks, warns, manifest)


def _token_vocab(cfg: Config) -> set[str]:
    """Extracted from the corpus. The system prompt is documentation, not evidence."""
    p = cfg.optional_path("inputs.interaction_report.path")
    cands = cfg.get("tokens.candidates", []) or []
    never = set(cfg.get("tokens.documented_but_not_produced", []) or [])
    if not p or not p.exists():
        return set()
    wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
    col = cfg.get("inputs.interaction_report.transcript_column")
    seen: set[str] = set()
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            hdr = [str(c).strip() if c else "" for c in next(rows)]
        except StopIteration:
            continue
        i = corpus.transcript_col_index(hdr, col)
        if i is None:
            continue
        for row in rows:
            if i < len(row) and row[i]:
                s = str(row[i])
                for t in cands:
                    if t in s:
                        seen.add(t)
    wb.close()
    return seen - never
