"""Transcript decoding, seed mining, and identifier harvesting.

The decoder registry is config-driven: `transcripts.decoders` lists the decoders
to try in order. Adding a client whose export format differs means adding a
decoder here once and naming it in that client's config — not editing a parser.
"""
from __future__ import annotations

import ast
import json
import re
from collections import Counter, defaultdict
from typing import Any, Callable

import openpyxl

from .config import Config

Turns = list[dict]
DECODERS: dict[str, Callable[[str, Config], Turns | None]] = {}


def decoder(name: str):
    def wrap(fn):
        DECODERS[name] = fn
        return fn
    return wrap


@decoder("pyrepr")
def _pyrepr(s: str, cfg: Config):
    # Single-quoted Python repr of a list of dicts. json.loads throws on this.
    if not (s.startswith("[") and "'speaker'" in s):
        return None
    return [{"speaker": d.get("speaker"), "text": d.get("text", "")}
            for d in ast.literal_eval(s)]


@decoder("json")
def _json(s: str, cfg: Config):
    if not (s.startswith("[") and '"speaker"' in s):
        return None
    return [{"speaker": d.get("speaker"), "text": d.get("text", "")} for d in json.loads(s)]


def _flat(name: str):
    @decoder(name)
    def _f(s: str, cfg: Config, _n=name):
        labels = cfg.get(f"transcripts.speaker_labels.{_n}", None)
        if not labels:
            return None
        a, c = labels["agent"], labels["customer"]
        if not re.search(rf"^\s*({re.escape(a)}|{re.escape(c)})", s):
            return None
        turns, spk = [], None
        for part in re.split(rf"({re.escape(a)}|{re.escape(c)})", s):
            p = part.strip()
            if p == a:
                spk = "agent"
            elif p == c:
                spk = "customer"
            elif spk and p:
                turns.append({"speaker": spk, "text": p})
        return turns
    return _f


for _n in ("flat_assistant_user", "flat_agent_customer", "flat_title_case"):
    _flat(_n)


def decode(raw: Any, cfg: Config) -> tuple[Turns | None, str]:
    """Returns (turns, decoder_name). decoder_name is EMPTY or UNRECOGNISED on failure."""
    if raw is None:
        return None, "EMPTY"
    s = str(raw).strip()
    if s.lower() in [m.lower() for m in cfg.get("transcripts.null_markers", [""])]:
        return None, "EMPTY"
    for name in cfg.get("transcripts.decoders"):
        fn = DECODERS.get(name)
        if not fn:
            continue
        try:
            out = fn(s, cfg)
        except Exception as e:  # a malformed row must not kill the run
            return None, f"{name}_FAIL:{type(e).__name__}"
        if out is not None:
            return out, name
    return None, "UNRECOGNISED"


def transcript_col_index(hdr: list[str], col: str) -> int | None:
    """Case-insensitive header match.

    Real workbooks capitalise inconsistently across sheets — the reference
    corpus has both `transcript` and `Transcript`. A case-sensitive match
    silently skips a whole sheet, which reads as "that encoding isn't in the
    corpus" rather than "we didn't look".
    """
    low = [h.strip().lower() for h in hdr]
    return low.index(col.strip().lower()) if col.strip().lower() in low else None


def decoder_census(cfg: Config) -> Counter:
    """Every transcript cell in the interaction report, by decoder."""
    wb = openpyxl.load_workbook(cfg.path("inputs.interaction_report.path"), data_only=True)
    col = cfg.get("inputs.interaction_report.transcript_column")
    out = Counter()
    for ws in wb.worksheets:
        hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
        i = transcript_col_index(hdr, col)
        if i is None:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            if i < len(row):
                out[decode(row[i], cfg)[1]] += 1
    return out


# ------------------------------------------------------- annotated errors

def mine_annotations(cfg: Config) -> list[dict]:
    """Rows a human corrected.

    Extracted as EVIDENCE OF CONFUSION. The human's wording is recorded as
    `human_claim`, never as ground truth — see docs/ARCHITECTURE.md.
    """
    wb = openpyxl.load_workbook(cfg.path("inputs.interaction_report.path"), data_only=True)
    prefixes = tuple(p.lower() for p in
                     cfg.get("inputs.interaction_report.annotation_column_prefixes"))
    tcol = cfg.get("inputs.interaction_report.transcript_column")
    lcols = cfg.get("inputs.interaction_report.label_columns", [])
    ccols = cfg.get("inputs.interaction_report.confidence_columns", [])
    nulls = {m.lower() for m in cfg.get("transcripts.null_markers", [])}
    rows = []
    for name in cfg.get("inputs.interaction_report.annotated_sheets"):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
        acols = [i for i, h in enumerate(hdr) if h.lower().startswith(prefixes)]
        if not acols:
            continue
        idx = {h: i for i, h in enumerate(hdr)}
        for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            vals = [row[i] for i in acols if i < len(row)]
            claim = [str(v).strip() for v in vals
                     if v and str(v).strip().lower() not in nulls]
            if not claim:
                continue
            g = lambda k: (row[idx[k]] if k in idx and idx[k] < len(row) else None)
            turns, dec = decode(g(tcol), cfg)
            rows.append({
                "source_sheet": name, "source_row": ri,
                "engine_label": next((g(c) for c in lcols if g(c)), None),
                "engine_confidence": next((g(c) for c in ccols if g(c)), None),
                "human_claim": " | ".join(claim),
                "decoder": dec, "turns": turns,
            })
    return rows


def empirical_pairs(cfg: Config, annotations: list[dict]) -> dict[str, list[dict]]:
    """Confusion the engine has actually demonstrated — the highest-weight tier."""
    rx = re.compile(cfg.get("taxonomy.code_regex"))
    out: dict[str, list[dict]] = defaultdict(list)
    for a in annotations:
        eng = str(a["engine_label"]).split(" - ")[0].strip() if a["engine_label"] else None
        for c in {m if isinstance(m, str) else m[0] for m in rx.findall(str(a["human_claim"]))}:
            if eng and c != eng:
                ev = f"production misfire {a['source_sheet']}#{a['source_row']}"
                out[eng].append({"other": c, "evidence": ev})
                out[c].append({"other": eng, "evidence": ev})
    return out


# ----------------------------------------------------------- redial seeds

def redial_seeds(cfg: Config) -> dict[str, list[dict]]:
    p = cfg.optional_path("inputs.redial_matrix.path")
    if not p or not p.exists():
        return {}
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb[cfg.get("inputs.redial_matrix.sheet")]
    hdr = [c.value for c in ws[1]]
    key = cfg.get("inputs.redial_matrix.code_column")
    ki = hdr.index(key)
    out: dict[str, list[dict]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if ki >= len(row) or not row[ki]:
            continue
        out[str(row[ki]).split(" - ")[0].strip()].append(
            {h: v for h, v in zip(hdr, row) if h and v})
    return out


# -------------------------------------------------- real-identifier harvest

def harvest_identifiers(cfg: Config) -> dict[str, set[str]]:
    """Every real identifier across every input workbook.

    Feeds validator V16 (PII containment). A generated case containing any of
    these is a release blocker, not a style issue.
    """
    spec = cfg.get("deidentify.harvest", {}) or {}
    found: dict[str, set[str]] = {k: set() for k in spec}
    paths = []
    for key in ("inputs.taxonomy.path", "inputs.redial_matrix.path",
                "inputs.interaction_report.path"):
        p = cfg.optional_path(key)
        if p and p.exists():
            paths.append(p)
    for ex in cfg.get("inputs.extras", []) or []:
        p = (cfg.root / ex["path"]).resolve()
        if p.exists():
            paths.append(p)

    for p in dict.fromkeys(paths):
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                hdr = [str(c).strip() if c else "" for c in next(rows)]
            except StopIteration:
                continue
            for kind, rule in spec.items():
                cols = [i for i, h in enumerate(hdr) if h in rule.get("columns", [])]
                if not cols:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    for i in cols:
                        if i >= len(row) or row[i] is None:
                            continue
                        s = str(row[i]).strip()
                        if not s or s.lower() in ("null", "none", "n/a", "0"):
                            continue
                        if rule.get("numeric") and not s.isdigit():
                            continue
                        if rule.get("length") and len(s) != rule["length"]:
                            continue
                        if len(s) < rule.get("min_length", 1):
                            continue
                        # A placeholder cell ("...", "--", "n.a.") harvested as a
                        # person name matches almost every file in the repo, so
                        # V16 reports a leak in the README and the real hits get
                        # lost in the noise. Require an actual word character.
                        if not rule.get("numeric") and not re.search(r"\w", s):
                            continue
                        found[kind].add(s)
        wb.close()
    return found
