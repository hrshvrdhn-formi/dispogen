"""Phase 7 — merge & render. Deterministic; openpyxl only.

Never let a model write a spreadsheet: formatting drift is guaranteed and silent.

The column contract is READ from the client's own output-format sheet at build
time, so a client that renames a column gets a workbook that still drops into
their pipeline. `render.contract_map` binds each contract column to a named
resolver below; an unmapped column renders blank rather than failing the build.
"""
from __future__ import annotations

import datetime as _dt
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import Config

HDR = PatternFill("solid", fgColor="1F3864")
CONTRACT_HDR = PatternFill("solid", fgColor="375623")
HDRF = Font(color="FFFFFF", bold=True, size=10)
SUBF = PatternFill("solid", fgColor="D9E2F3")
AMBF = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

RESOLVERS = {}


def resolver(name):
    def wrap(fn):
        RESOLVERS[name] = fn
        return fn
    return wrap


@resolver("sn")
def _sn(c, d, cfg): return c.get("sn")


@resolver("group_code")
def _gc(c, d, cfg): return d["group"].split(" - ")[0].strip()


@resolver("sub_name")
def _sn2(c, d, cfg):
    parts = d["sub"].split(" - ", 1)
    return parts[1] if len(parts) > 1 else d["sub"]


@resolver("engine_code")
def _ec(c, d, cfg): return d["engine_code"]


@resolver("probe_type_label")
def _pt(c, d, cfg):
    return cfg.get("inputs.output_format.probe_type_values")[c["probe_type"]]


@resolver("scenario")
def _sc(c, d, cfg): return c.get("scenario", "")


@resolver("transcript_text")
def _tt(c, d, cfg):
    return "\n".join(f"{t['speaker'].upper()}: {t['text']}" for t in c.get("transcript", []))


@resolver("transcript_json")
def _tj(c, d, cfg):
    return json.dumps(c.get("transcript", []), ensure_ascii=False)


@resolver("redial_required")
def _rr(c, d, cfg): return c.get("redial", {}).get("is_required", "")


@resolver("anchor")
def _an(c, d, cfg): return c.get("redial", {}).get("anchor_date", d.get("anchor", ""))


@resolver("redial_schedule")
def _rs(c, d, cfg): return c.get("redial", {}).get("schedule", "")


@resolver("blank")
def _bl(c, d, cfg): return ""


APPENDED = {
    "test_case_id": lambda c, d: c.get("test_case_id", ""),
    "probe_type": lambda c, d: c.get("probe_type", ""),
    "archetype": lambda c, d: c.get("archetype", ""),
    "certification_grade": lambda c, d: c.get("declared_grade", ""),
    "expected_group": lambda c, d: c.get("expected_group", ""),
    "expected_sub": lambda c, d: c.get("expected_sub") or "",
    "expected_expanded": lambda c, d: c.get("expected_expanded") or "",
    "must_not_select": lambda c, d: ", ".join(c.get("must_not_select", [])),
    "rival_code": lambda c, d: c.get("rival_code") or "",
    "trap_phrase": lambda c, d: c.get("trap_phrase") or "",
    "decisive_evidence": lambda c, d: c.get("decisive_evidence", ""),
    "cited_clause": lambda c, d: c.get("cited_clause", ""),
    "rebutted_rivals": lambda c, d: " | ".join(
        f"{r['code']}: {r['why']}" for r in c.get("rebutted_rivals", [])),
    "precedence_rule_applied": lambda c, d: c.get("precedence_rule_applied") or "",
    "source_of_truth_class": lambda c, d: d.get("source_of_truth_class", ""),
    "pre_call_parameters": lambda c, d: json.dumps(
        c.get("pre_call_parameters", {}), ensure_ascii=False),
    "perturbations": lambda c, d: ", ".join(c.get("perturbations", [])),
    "redial_context": lambda c, d: c.get("redial", {}).get("context", ""),
    "redial_basis": lambda c, d: c.get("redial", {}).get("basis", ""),
    "certification_status": lambda c, d: c.get("certification_status", "PROVISIONAL"),
    "generated_by": lambda c, d: d.get("generated_by", ""),
}


def _style_header(ws, ncols, contract_n=0):
    for i in range(1, ncols + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = CONTRACT_HDR if i <= contract_n else HDR
        cell.font = HDRF
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"


def _widths(ws, spec):
    for i, w in enumerate(spec, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def contract_columns(cfg: Config) -> list[str]:
    wb = openpyxl.load_workbook(cfg.path("inputs.output_format.path"), data_only=True)
    ws = wb[cfg.get("inputs.output_format.sheet")]
    return [str(c.value).strip() if c.value else "" for c in ws[1]]


def build(cfg: Config, docs: list[dict], register: dict, manifest: dict,
          prescan_findings: list[dict], gates: list[list], out_path: Path) -> Path:
    contract = contract_columns(cfg)
    cmap = cfg.get("render.contract_map", {}) or {}
    appended = cfg.get("render.appended_columns")
    wb = openpyxl.Workbook()

    # ---------------- Sheet 1 : Test Cases -------------------------------
    ws = wb.active
    ws.title = "Test Cases"
    ws.append(contract + appended)
    _style_header(ws, len(contract) + len(appended), len(contract))
    for d in docs:
        for c in d["cases"]:
            row = []
            for col in contract:
                fn = RESOLVERS.get(cmap.get(col, "blank"), RESOLVERS["blank"])
                row.append(fn(c, d, cfg))
            row += [APPENDED[a](c, d) if a in APPENDED else "" for a in appended]
            ws.append(row)
    _widths(ws, [5, 8, 22, 26, 38, 52, 90, 24, 22, 30, 40] +
            [30, 9, 24, 13, 26, 28, 30, 22, 10, 26, 46, 52, 70, 12, 16, 46, 46, 66, 46, 46, 26])
    grade_col = len(contract) + appended.index("certification_grade") + 1
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 30
        for cix in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=cix)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN
        if ws.cell(row=r, column=grade_col).value in ("SUB", "GROUP"):
            for cix in range(1, len(contract) + 1):
                ws.cell(row=r, column=cix).fill = SUBF

    # ---------------- Sheet 2 : Ambiguous Scenarios ----------------------
    ws2 = wb.create_sheet("Ambiguous Scenarios")
    ws2.append(["Ambiguity ID", "Class", "Class Name", "Subject",
                "Competing Label A", "Clause A", "Competing Label B", "Clause B",
                "Why Unresolved", "Evidence", "Production Frequency",
                "Proposed Amendment", "Blast Radius", "Resolution"])
    _style_header(ws2, 14)
    for e in register.get("entries", []):
        comp = list(e.get("competing", [])) + [{"code": "", "clause": ""}] * 2
        ws2.append([e["id"], e["class"], e.get("class_name", ""), e["subject"],
                    comp[0]["code"], comp[0]["clause"], comp[1]["code"], comp[1]["clause"],
                    e.get("why_unresolved", ""), e.get("evidence", ""),
                    e.get("production_frequency", ""), e.get("proposed_amendment", ""),
                    e.get("blast_radius", ""),
                    e.get("resolution", "OPEN - needs taxonomy owner decision")])
    _widths(ws2, [20, 7, 30, 60, 20, 62, 20, 62, 90, 52, 46, 80, 42, 40])
    for r in range(2, ws2.max_row + 1):
        ws2.row_dimensions[r].height = 92
        for cix in range(1, 15):
            ws2.cell(row=r, column=cix).alignment = Alignment(vertical="top", wrap_text=True)
            ws2.cell(row=r, column=cix).border = THIN
        ws2.cell(row=r, column=2).fill = AMBF

    # ---------------- Sheet 3 : Coverage Matrix --------------------------
    ws3 = wb.create_sheet("Coverage Matrix")
    ws3.append(["Engine Code", "Label", "Status", "Source of Truth", "FN", "FP", "Total",
                "@EXPANDED", "@SUB", "@GROUP", "Rivals Probed", "Rival Tiers",
                "Perturbation Axes", "Redial Seeds", "Shortfall", "Allocation Notes"])
    _style_header(ws3, 16)
    # Every leaf in the taxonomy gets a row, not just the ones with cases. A
    # coverage matrix that lists only what was authored reports 100% coverage of
    # itself and says nothing about the dispositions still outstanding.
    authored = {d["engine_code"]: d for d in docs}
    for pk in sorted((cfg.root / "compiled" / "packs").glob("*.json")):
        pack = json.loads(pk.read_text(encoding="utf-8"))
        code = pack["engine_code"]
        alloc = pack["quota"]["fp_allocation"]
        common = [", ".join(a["rival_num"] for a in alloc),
                  ", ".join(sorted({a["tier"] for a in alloc}))]
        tail = [len(pack.get("redial_seeds", [])), pack["quota"]["shortfall"],
                "; ".join(pack["quota"].get("allocation_notes", []))]
        d = authored.get(code)
        if d is None:
            ws3.append([code, pack["label"], "PENDING", pack["source_of_truth_class"],
                        0, 0, 0, 0, 0, 0] + common + [""] + tail)
            continue
        cs = d["cases"]
        g = Counter(c.get("declared_grade") for c in cs)
        axes = {p.split(":")[0] for c in cs for p in c.get("perturbations", [])}
        status = sorted({c.get("certification_status", "PROVISIONAL") for c in cs})
        ws3.append([code, d["label"], "/".join(status), d.get("source_of_truth_class", ""),
                    sum(1 for c in cs if c["probe_type"] == "FN"),
                    sum(1 for c in cs if c["probe_type"] == "FP"), len(cs),
                    g["EXPANDED"], g["SUB"], g["GROUP"]] + common +
                   [", ".join(sorted(axes))] + tail)
    _widths(ws3, [26, 32, 14, 16, 5, 5, 7, 12, 9, 11, 22, 30, 62, 14, 10, 60])
    for r in range(2, ws3.max_row + 1):
        ws3.row_dimensions[r].height = 46
        for cix in range(1, 17):
            ws3.cell(row=r, column=cix).alignment = Alignment(vertical="top", wrap_text=True)
        if ws3.cell(row=r, column=3).value == "PENDING":
            ws3.cell(row=r, column=3).fill = AMBF

    # ---------------- Sheet 4 : Re-Dial Expectations ---------------------
    ws4 = wb.create_sheet("Re-Dial Expectations")
    win = cfg.get("inputs.redial_matrix.calling_window", {"start": "09:00", "end": "21:00"})
    ws4.append(["test_case_id", "Expected Disposition (graded)", "Grade", "Is Redial Required",
                "Anchor", "Next Interaction", f"Inside {win['start']}-{win['end']}",
                "After Anchor", "Context for the Next Interaction", "Basis"])
    _style_header(ws4, 10)
    tre = re.compile(r"\d{2}:\d{2}")
    for d in docs:
        for c in d["cases"]:
            rd = c.get("redial", {})
            m = tre.search(str(rd.get("schedule", "")))
            inside = ""
            if m:
                h = int(m.group(0)[:2])
                inside = "YES" if int(win["start"][:2]) <= h < int(win["end"][:2]) else "NO"
            graded = (c.get("expected_expanded") or c.get("expected_sub")
                      or c.get("expected_group"))
            ws4.append([c["test_case_id"], graded, c.get("declared_grade"),
                        rd.get("is_required", ""), rd.get("anchor_date", ""),
                        rd.get("schedule", ""), inside, "YES" if m else "n/a",
                        rd.get("context", ""), rd.get("basis", "")])
    _widths(ws4, [32, 40, 12, 26, 22, 50, 16, 14, 88, 76])
    for r in range(2, ws4.max_row + 1):
        ws4.row_dimensions[r].height = 56
        for cix in range(1, 11):
            ws4.cell(row=r, column=cix).alignment = Alignment(vertical="top", wrap_text=True)

    # ---------------- Sheet 5 : Certification Log ------------------------
    ws5 = wb.create_sheet("Certification Log")
    ws5.append(["Gate", "What it checks", "Mechanism", "Status", "Result", "Notes"])
    _style_header(ws5, 6)
    for g in gates:
        ws5.append(g)
    _widths(ws5, [26, 62, 46, 12, 34, 76])
    for r in range(2, ws5.max_row + 1):
        ws5.row_dimensions[r].height = 40
        for cix in range(1, 7):
            ws5.cell(row=r, column=cix).alignment = Alignment(vertical="top", wrap_text=True)
        st = ws5.cell(row=r, column=4).value
        ws5.cell(row=r, column=4).fill = PatternFill(
            "solid", fgColor="C6EFCE" if st == "RUN" else "FFC7CE")

    # ---------------- Sheet 6 : Taxonomy Defects -------------------------
    ws6 = wb.create_sheet("Taxonomy Defects")
    ws6.append(["Class", "Class Name", "Subject", "Detail", "Impact"])
    _style_header(ws6, 5)
    for f in sorted(prescan_findings, key=lambda x: x["class"]):
        ws6.append([f["class"], f["class_name"], str(f["subject"]),
                    f["detail"][:600], f["impact"]])
    _widths(ws6, [7, 34, 26, 110, 70])
    for r in range(2, ws6.max_row + 1):
        ws6.row_dimensions[r].height = 44
        for cix in range(1, 6):
            ws6.cell(row=r, column=cix).alignment = Alignment(vertical="top", wrap_text=True)

    # ---------------- Sheet 7 : Run Metadata -----------------------------
    ws7 = wb.create_sheet("Run Metadata")
    ws7.append(["Key", "Value"])
    _style_header(ws7, 2)
    rows = [("client", manifest.get("client")),
            ("config_hash", manifest.get("config_hash")),
            ("preflight", manifest.get("preflight")),
            ("taxonomy", json.dumps(manifest.get("counts", {}))),
            ("transcript_decoders", json.dumps(manifest.get("decoders", {}))),
            ("token_vocabulary", ", ".join(manifest.get("token_vocabulary", []))),
            ("dispositions_rendered", len(docs)),
            ("cases_rendered", sum(len(d["cases"]) for d in docs)),
            ("ambiguity_register_entries", len(register.get("entries", []))),
            ("taxonomy_defect_findings", len(prescan_findings)),
            ("built_at", _dt.datetime.now().strftime("%Y-%m-%d %H:%M"))]
    for k, v in rows:
        ws7.append([k, v])
    for w in manifest.get("warnings", []):
        ws7.append(["warning", w])
    _widths(ws7, [42, 112])
    for r in range(2, ws7.max_row + 1):
        for cix in (1, 2):
            ws7.cell(row=r, column=cix).alignment = Alignment(vertical="top", wrap_text=True)
        ws7.cell(row=r, column=1).font = Font(bold=True)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
