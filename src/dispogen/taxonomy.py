"""Taxonomy normalisation, confusion graph, precedence ladder, rival allocation.

All config-driven. Every structure is keyed on engine_code — numeric labels are
display strings only, because they are not unique across levels in every client
taxonomy (and are not in the reference one).
"""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict
from typing import Any

import openpyxl

from .config import Config


def norm_label(v: Any) -> str:
    return str(v).split(" - ")[0].strip()


@dataclass
class Leaf:
    engine_code: str
    label: str
    num: str
    group: str
    group_code: str
    sub: str
    sub_num: str
    description: str = ""
    decision_rules: str = ""
    engineering_note: str = ""
    source_of_truth_class: str = "transcript"

    def dict(self) -> dict:
        return asdict(self)


@dataclass
class Taxonomy:
    leaves: list[Leaf]
    subs: dict[tuple[str, str], str]
    groups: dict[str, str]
    corpus: str = ""

    by_code: dict[str, Leaf] = field(default_factory=dict)
    by_num: dict[str, Leaf] = field(default_factory=dict)

    def __post_init__(self):
        self.by_code = {l.engine_code: l for l in self.leaves}
        self.by_num = {l.num: l for l in self.leaves}

    def siblings(self, leaf: Leaf) -> list[Leaf]:
        return [l for l in self.leaves
                if l.group == leaf.group and l.sub == leaf.sub and l.num != leaf.num]

    def sub_nums(self) -> set[str]:
        return {norm_label(s) for (_, s) in self.subs}

    def sub_owners(self, num: str) -> list[str]:
        return [g for (g, s) in self.subs if norm_label(s) == num]


# ------------------------------------------------------------------ loading

def load(cfg: Config) -> Taxonomy:
    C = cfg.get("inputs.taxonomy.columns")
    ff = set(cfg.get("inputs.taxonomy.forward_fill", ["group", "group_desc", "sub", "sub_desc"]))
    wb = openpyxl.load_workbook(cfg.path("inputs.taxonomy.path"), data_only=True)
    sheet = cfg.get("inputs.taxonomy.sheet")
    if sheet not in wb.sheetnames:
        raise ValueError(f"sheet {sheet!r} not in {wb.sheetnames}")
    ws = wb[sheet]

    hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {}
    for key, colname in C.items():
        if colname not in hdr:
            raise ValueError(
                f"column {colname!r} (config key inputs.taxonomy.columns.{key}) "
                f"not found. Sheet has: {hdr}")
        idx[key] = hdr.index(colname)

    def cell(row, key, dflt=""):
        i = idx.get(key)
        return "" if i is None or i >= len(row) or row[i] is None else str(row[i]).strip()

    leaves, subs, groups = [], {}, {}
    g = s = None
    corpus_parts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        corpus_parts += [str(v) for v in row if v]
        if cell(row, "group"):
            g = cell(row, "group")
            groups[g] = cell(row, "group_desc")
        elif "group" not in ff:
            g = cell(row, "group") or g
        if cell(row, "sub"):
            s = cell(row, "sub")
            subs[(g, s)] = cell(row, "sub_desc")
        elif "sub" not in ff:
            s = cell(row, "sub") or s
        exp = cell(row, "expanded")
        if not exp:
            continue
        leaves.append(Leaf(
            engine_code=cell(row, "engine_code"), label=exp, num=norm_label(exp),
            group=g or "", group_code=norm_label(g or ""), sub=s or "",
            sub_num=norm_label(s or ""),
            description=cell(row, "expanded_desc"),
            decision_rules=cell(row, "decision_rules"),
            engineering_note=cell(row, "engineering_note"),
        ))

    tax = Taxonomy(leaves, subs, groups, corpus="\n".join(corpus_parts))
    for l in tax.leaves:
        l.source_of_truth_class = classify_source(cfg, l)
    return tax


def classify_source(cfg: Config, leaf: Leaf) -> str:
    ov = cfg.get("taxonomy.source_of_truth_overrides", {}) or {}
    if leaf.engine_code in ov:
        return ov[leaf.engine_code]
    hay = (leaf.decision_rules + " " + leaf.engineering_note).lower()
    for rule in cfg.get("taxonomy.source_of_truth_rules", []) or []:
        if rule["match"].lower() in hay:
            return rule["class"]
    return cfg.get("taxonomy.source_of_truth_default", "transcript")


# ---------------------------------------------------------- confusion graph

def confusion_graph(cfg: Config, tax: Taxonomy) -> dict:
    rx = re.compile(cfg.get("taxonomy.code_regex"))
    sub_nums = tax.sub_nums()
    graph = {}
    for l in tax.leaves:
        rivals, dangling = [], []
        for m in dict.fromkeys(rx.findall(l.decision_rules)):
            m = m if isinstance(m, str) else m[0]
            if m == l.num:
                continue
            if m in tax.by_num:
                rivals.append({"num": m, "engine_code": tax.by_num[m].engine_code,
                               "level": "expanded", "source": "explicit"})
            elif m in sub_nums:
                # A sub label can exist under more than one group. Resolve to
                # this leaf's own group first; flag when we cannot.
                owners = tax.sub_owners(m)
                same = [o for o in owners if o == l.group]
                rivals.append({"num": m, "engine_code": f"SUB::{m}", "level": "sub",
                               "source": "explicit",
                               "ambiguous_parent": len(owners) > 1 and not same})
            else:
                dangling.append(m)
        graph[l.engine_code] = {"num": l.num, "rivals": rivals, "dangling": dangling}
    return graph


# -------------------------------------------------------- precedence ladder

def precedence(cfg: Config, tax: Taxonomy) -> tuple[list[dict], list[str]]:
    """Returns (ladder, missing_anchor_ids).

    An anchor that no longer resolves means the source document changed. That is
    a preflight failure, not a warning — the ladder would silently stop matching.
    """
    ladder = list(cfg.get("precedence"))
    missing = [p["id"] for p in ladder if p["anchor"] not in tax.corpus]
    return ladder, missing


# ------------------------------------------------------- rival allocation

def allocate(cfg: Config, tax: Taxonomy, leaf: Leaf, graph: dict,
             empirical: dict[str, list[dict]]) -> dict:
    """Fill FP slots by semantic role, per config.quota.fp_slots.

    Rank-ordering the candidate pool and zipping it onto the archetypes drops the
    highest-information roles on any leaf with a crowded sibling set, while still
    reporting the quota as filled. Slots therefore reserve their own pool.
    """
    tiers = cfg.get("tiers")
    kids = tax.siblings(leaf)
    singleton_sub = not kids

    def mk(num, code, tier, why, level="expanded"):
        return {"num": num, "engine_code": code, "tier": tier,
                "weight": tiers[tier]["weight"], "level": level, "why": why}

    pools: dict[str, list[dict]] = defaultdict(list)
    for e in empirical.get(leaf.num, []):
        t = tax.by_num.get(e["other"])
        pools["empirical"].append(
            mk(e["other"], t.engine_code if t else f"NUM::{e['other']}", "empirical", e["evidence"]))
    for r in graph[leaf.engine_code]["rivals"]:
        if r["level"] == "expanded":
            pools["explicit"].append(mk(r["num"], r["engine_code"], "explicit", tiers["explicit"]["why"]))
    for k in kids:
        pools["sibling"].append(mk(k.num, k.engine_code, "sibling", tiers["sibling"]["why"]))
    for l in tax.leaves:
        if l.group != leaf.group and l.num in {r["num"] for r in graph[leaf.engine_code]["rivals"]}:
            pools["out_of_class"].append(mk(l.num, l.engine_code, "out_of_class", tiers["out_of_class"]["why"]))
    if leaf.source_of_truth_class == "transcript":
        for num in cfg.get("taxonomy.speaker_identity_rivals", []) or []:
            t = tax.by_num.get(num)
            if t and t.group != leaf.group and num != leaf.num:
                pools["out_of_class"].append(
                    mk(num, t.engine_code, "out_of_class",
                       "a non-policyholder voices the claim -- third-party boundary"))
    if not singleton_sub:
        pools["parent"] = [mk(leaf.sub_num, f"SUB::{leaf.sub_num}", "parent",
                              tiers["parent"]["why"], "sub")]
    pools["group"] = [mk(leaf.group_code, f"GROUP::{leaf.group_code}", "group",
                         tiers["group"]["why"], "group")]

    slots, used, notes = {}, {leaf.num}, []
    if singleton_sub:
        notes.append("stop-at-parent is degenerate: this sub has exactly one child")

    for spec in cfg.get("quota.fp_slots"):
        chosen = None
        for pool_names in (spec.get("pools", []), spec.get("fallback_pools", [])):
            for pname in pool_names:
                for cand in sorted(pools.get(pname, []), key=lambda x: -x["weight"]):
                    if cand["num"] not in used:
                        chosen = cand
                        break
                if chosen:
                    break
            if chosen:
                break
            if pool_names is spec.get("pools", []) and spec.get("fallback_pools"):
                notes.append(f"{spec['slot']} degraded: no {spec['pools']} candidate, using fallback")
        if chosen:
            used.add(chosen["num"])
            slots[spec["slot"]] = {**chosen, "role": spec["role"]}
        else:
            notes.append(f"{spec['slot']} UNFILLED ({spec['role']})")

    ordered = [slots[s["slot"]] for s in cfg.get("quota.fp_slots") if s["slot"] in slots]
    return {"rivals": ordered, "slots": slots, "notes": notes,
            "shortfall": len(cfg.get("quota.fp_slots")) - len(ordered),
            "singleton_sub": singleton_sub}


def integrity(tax: Taxonomy) -> dict:
    codes = [l.engine_code for l in tax.leaves]
    trips = [(l.group, l.sub, l.label) for l in tax.leaves]
    return {
        "leaves": len(tax.leaves), "groups": len(tax.groups), "subs": len(tax.subs),
        "engine_codes_unique": len(set(codes)) == len(codes),
        "engine_code_duplicates": [k for k, v in Counter(codes).items() if v > 1],
        "triple_unique": len(set(trips)) == len(trips),
        "empty_rules": [l.engine_code for l in tax.leaves if not l.decision_rules.strip()],
        "empty_codes": [l.label for l in tax.leaves if not l.engine_code],
    }
